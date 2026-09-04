from __future__ import annotations

import json
import re
from dataclasses import dataclass, replace
from pathlib import Path

import pytest
from openpyxl import load_workbook

from llm_abm_sim import ConcurrentRobustnessStudy
from llm_abm_sim import concurrent_robustness_report as report_module
from llm_abm_sim.concurrent_robustness_v2_evidence import (
    _read_closed_concurrent_robustness_v2_report_source,
)
from llm_abm_sim.concurrent_robustness_v2_report import (
    ConcurrentRobustnessV2ReportError,
    _provider_audit_rows,
)
from tests.integration.test_concurrent_message_experiment_runner import _make_validation_report_source
from tests.integration.test_concurrent_robustness_v2 import _v2_adapters, _v2_manifest
from tests.integration.test_full_pool_presentation_bundle import (
    _formal_realized_full_pool_source,
    _historical_candidate,
)


def _snapshot(root: Path) -> dict[str, bytes]:
    return {
        path.relative_to(root).as_posix(): path.read_bytes()
        for path in sorted(root.rglob("*"))
        if path.is_file()
    }


def _closed_v2_study(root: Path) -> Path:
    source = _make_validation_report_source(root, "v2-report-source")
    manifest = _v2_manifest(source, output_identity="v2-report-study")
    adapters, _ = _v2_adapters(manifest)
    result = ConcurrentRobustnessStudy().run(
        manifest,
        adapters,
        root / "v2-report-workspace",
    )
    assert result.study_root is not None
    return result.study_root


@dataclass(frozen=True)
class _ReportFixture:
    full_pool: Path
    full_pool_manifest_sha256: str
    historical_formal: Path
    historical_study: Path
    historical_candidate: Path
    v2_study: Path
    destination: Path
    before: dict[Path, dict[str, bytes]]


@pytest.fixture
def v2_report_fixture(tmp_path: Path, monkeypatch) -> _ReportFixture:
    full_pool, full_pool_manifest_sha256 = _formal_realized_full_pool_source(
        tmp_path / "full-pool",
        monkeypatch,
    )
    historical_formal, historical_study, historical_candidate = _historical_candidate(
        tmp_path / "historical"
    )
    v2_study = _closed_v2_study(tmp_path / "v2")
    destination = tmp_path / "v2-realized-candidate"
    protected = (full_pool, historical_formal, historical_study, historical_candidate, v2_study)
    before = {root: _snapshot(root) for root in protected}
    created = report_module._REPORT_PRESENTATION.compose_v2_realized_candidate(
        full_pool_source_root=full_pool,
        full_pool_manifest_sha256=full_pool_manifest_sha256,
        historical_formal_root=historical_formal,
        historical_study_root=historical_study,
        historical_candidate_dir=historical_candidate,
        v2_study_root=v2_study,
        destination_dir=destination,
    )
    assert created == destination.resolve()
    return _ReportFixture(
        full_pool=full_pool,
        full_pool_manifest_sha256=full_pool_manifest_sha256,
        historical_formal=historical_formal,
        historical_study=historical_study,
        historical_candidate=historical_candidate,
        v2_study=v2_study,
        destination=destination,
        before=before,
    )


def test_report_interface_composes_one_table_first_v2_candidate_and_deterministic_workbook(
    v2_report_fixture: _ReportFixture,
) -> None:
    fixture = v2_report_fixture
    full_pool = fixture.full_pool
    full_pool_manifest_sha256 = fixture.full_pool_manifest_sha256
    historical_formal = fixture.historical_formal
    historical_study = fixture.historical_study
    historical_candidate = fixture.historical_candidate
    v2_study = fixture.v2_study
    destination = fixture.destination
    before = fixture.before
    protected = tuple(before)
    created = destination.resolve()

    assert created == destination.resolve()
    assert all(_snapshot(root) == before[root] for root in protected)
    expected_new_files = {
        "artifact_manifest.json",
        "robustness_v2_report_payload.json",
        "robustness_v2_projection.json",
        "robustness_v2_realized_main.csv",
        "robustness_v2_judgment_audit.csv",
        "robustness_v2_prompt_catalog.json",
        "robustness_v2_provider_audit.csv",
        "robustness_v2_cell_batch_evidence.csv",
        "robustness_v2_teacher_results.xlsx",
        "prompt-model-realized-mechanism.mmd",
    }
    assert expected_new_files <= {path.name for path in destination.iterdir() if path.is_file()}

    payload = json.loads((destination / "robustness_v2_report_payload.json").read_text(encoding="utf-8"))
    projection = json.loads((destination / "robustness_v2_projection.json").read_text(encoding="utf-8"))
    manifest = json.loads((destination / "artifact_manifest.json").read_text(encoding="utf-8"))
    assert payload["production_deploy_eligible"] is False
    assert payload["provider_calls_during_composition"] == 0
    assert payload["canonical_deployment_triggered"] is False
    assert projection["primary_outcome"] == "abm_realized_engagement"
    assert len(projection["realized_main_rows"]) == 180
    assert len(projection["judgment_audit_rows"]) == 180
    assert len(projection["prompt_catalog"]) == 4
    assert len(projection["provider_audit_rows"]) == 5
    assert all(
        row["execution_profile"] == "deterministic_validation"
        and row["condition_evidence_scope"]
        == "planned_condition_not_executed_in_deterministic_validation"
        and json.loads(row["observed_provider_route_counts"])
        == {"injected_deterministic_validation": row["physical_attempt_count"]}
        and row["planned_provider_route"] != "injected_deterministic_validation"
        for row in projection["provider_audit_rows"]
    )
    assert len(projection["cell_batch_evidence_rows"]) == 40
    assert manifest["production_deploy_eligible"] is False
    assert manifest["provider_calls_during_composition"] == 0

    workbook_path = destination / "robustness_v2_teacher_results.xlsx"
    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    assert workbook.sheetnames == [
        "README & Lineage",
        "Realized Main",
        "Judgment Audit",
        "Prompt Catalog",
        "Provider Audit",
        "Cell-Batch Evidence",
    ]
    assert all(workbook[name].freeze_panes == "A2" for name in workbook.sheetnames)
    realized = workbook["Realized Main"]
    assert [cell.value for cell in realized[1]] == [
        "Model",
        "Prompt",
        "Segment",
        "Message",
        "Like count",
        "Comment count",
        "Share count",
        "Engagement count",
        "Exposure count",
        "Engagement rate",
        "Prompt anchor",
    ]
    first = projection["realized_main_rows"][0]
    assert [cell.value for cell in realized[2]] == [
        first["model"],
        first["prompt"],
        first["segment"],
        first["message"],
        first["like_count"],
        first["comment_count"],
        first["share_count"],
        first["engagement_count"],
        first["exposure_count"],
        first["engagement_rate"],
        first["prompt_anchor"],
    ]
    workbook.close()

    report = (destination / "report.html").read_text(encoding="utf-8")
    assert 'data-testid="robustness-v2-report"' in report
    assert 'data-v2-active-view="realized"' in report
    assert 'data-testid="robustness-v2-realized-view"' in report
    assert 'data-testid="robustness-v2-judgment-view"' in report
    assert 'data-testid="robustness-v2-model-select"' in report
    assert 'data-testid="robustness-v2-prompt-select"' in report
    assert 'data-testid="robustness-v2-prompt-catalog"' in report
    assert 'data-testid="robustness-v2-mechanism-svg"' in report
    assert 'data-testid="robustness-v2-mechanism-fallback"' in report
    assert "Antigravity OpenAI-compatible gateway" in report
    assert "Planned Provider condition" in report
    assert "injected_deterministic_validation" in report
    assert "not direct Gemini Developer API" in report
    assert "不可见上下文" in report
    assert "client-submitted Prompt" in report
    assert "Historical 16-cell Judgment Reference" in report
    assert "realized_reason" not in report
    assert report.index('data-testid="robustness-v2-realized-view"') < report.index(
        'data-testid="full-pool-main-experiment"'
    )

    linked = set(re.findall(r'<a\b[^>]*\bdownload\b[^>]*\bhref="([^"]+)"', report))
    assert linked == set(manifest["approved_downloads"].values())
    assert all((destination / relative).is_file() for relative in linked)
    assert _snapshot(destination / "historical-1000") == before[historical_candidate]
    assert _snapshot(destination / "full-pool-source") == before[full_pool]

    report_module._REPORT_PRESENTATION.validate_v2_realized_candidate(
        destination,
        full_pool_source_root=full_pool,
        full_pool_manifest_sha256=full_pool_manifest_sha256,
        historical_formal_root=historical_formal,
        historical_study_root=historical_study,
        historical_candidate_dir=historical_candidate,
        v2_study_root=v2_study,
    )


def _validate_fixture(fixture: _ReportFixture, candidate: Path | None = None) -> None:
    report_module._REPORT_PRESENTATION.validate_v2_realized_candidate(
        candidate or fixture.destination,
        full_pool_source_root=fixture.full_pool,
        full_pool_manifest_sha256=fixture.full_pool_manifest_sha256,
        historical_formal_root=fixture.historical_formal,
        historical_study_root=fixture.historical_study,
        historical_candidate_dir=fixture.historical_candidate,
        v2_study_root=fixture.v2_study,
    )


def test_v2_report_candidate_repeats_byte_for_byte(
    v2_report_fixture: _ReportFixture,
) -> None:
    fixture = v2_report_fixture
    repeated = fixture.destination.parent / "v2-realized-candidate-repeat"
    report_module._REPORT_PRESENTATION.compose_v2_realized_candidate(
        full_pool_source_root=fixture.full_pool,
        full_pool_manifest_sha256=fixture.full_pool_manifest_sha256,
        historical_formal_root=fixture.historical_formal,
        historical_study_root=fixture.historical_study,
        historical_candidate_dir=fixture.historical_candidate,
        v2_study_root=fixture.v2_study,
        destination_dir=repeated,
    )

    assert _snapshot(repeated) == _snapshot(fixture.destination)
    _validate_fixture(fixture, repeated)


def test_v2_report_composition_removes_installed_candidate_when_final_source_check_fails(
    v2_report_fixture: _ReportFixture,
    monkeypatch,
) -> None:
    fixture = v2_report_fixture
    destination = fixture.destination.parent / "v2-realized-source-drift"

    def reject_mutated_source(_inputs: object) -> None:
        raise report_module._RobustnessReportClosureError("source changed during composition")

    monkeypatch.setattr(
        report_module._REPORT_PRESENTATION,
        "_assert_v2_realized_report_inputs_unchanged",
        reject_mutated_source,
    )
    with pytest.raises(report_module._RobustnessReportClosureError, match="source changed"):
        report_module._REPORT_PRESENTATION.compose_v2_realized_candidate(
            full_pool_source_root=fixture.full_pool,
            full_pool_manifest_sha256=fixture.full_pool_manifest_sha256,
            historical_formal_root=fixture.historical_formal,
            historical_study_root=fixture.historical_study,
            historical_candidate_dir=fixture.historical_candidate,
            v2_study_root=fixture.v2_study,
            destination_dir=destination,
        )

    assert not destination.exists()


def test_provider_audit_rejects_validation_routes_labeled_as_formal(
    v2_report_fixture: _ReportFixture,
) -> None:
    report_source = _read_closed_concurrent_robustness_v2_report_source(
        v2_report_fixture.v2_study
    )
    crossed = replace(
        report_source,
        manifest=report_source.manifest.model_copy(
            update={"execution_profile": "formal"}
        ),
    )

    with pytest.raises(
        ConcurrentRobustnessV2ReportError,
        match="Formal Provider route or billing evidence",
    ):
        _provider_audit_rows(crossed)


def test_v2_report_validator_rejects_tampered_projection_workbook_report_and_lineage_copy(
    v2_report_fixture: _ReportFixture,
) -> None:
    fixture = v2_report_fixture
    targets = (
        fixture.destination / "robustness_v2_projection.json",
        fixture.destination / "robustness_v2_teacher_results.xlsx",
        fixture.destination / "report.html",
        fixture.destination / "historical-1000" / "report.html",
        fixture.destination / "artifact_manifest.json",
    )
    for target in targets:
        original = target.read_bytes()
        target.write_bytes(original + b"\n")
        with pytest.raises(
            report_module._RobustnessReportClosureError,
            match="failed independent validation",
        ):
            _validate_fixture(fixture)
        target.write_bytes(original)

    _validate_fixture(fixture)
