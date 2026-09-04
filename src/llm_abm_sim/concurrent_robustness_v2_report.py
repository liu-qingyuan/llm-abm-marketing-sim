from __future__ import annotations

import csv
import hashlib
import html
import io
import json
import math
import os
import re
import shutil
import tempfile
import zipfile
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path, PurePosixPath
from typing import Any, cast

import xlsxwriter
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .concurrent_message_mechanism_presentation import _MECHANISM_PRESENTATION
from .concurrent_robustness_v2 import _V2_MODELS
from .concurrent_robustness_v2_evidence import (
    _assert_concurrent_robustness_v2_report_source_unchanged,
    _ConcurrentRobustnessV2ReportSource,
)
from .full_pool_presentation import (
    compose_full_pool_presentation_bundle,
    validate_full_pool_presentation_bundle,
)
from .full_pool_two_stage_replay import ClosedFullPoolTwoStageSource
from .prompt_contracts import CONCURRENT_ROBUSTNESS_PROMPT_REGISTRY
from .providers.robustness import robustness_provider_disclosures

_REPORT_SCHEMA = "concurrent-robustness-v2-report-payload-v1"
_PROJECTION_SCHEMA = "concurrent-robustness-v2-report-projection-v1"
_CANDIDATE_MANIFEST_SCHEMA = "concurrent-robustness-v2-report-candidate-manifest-v1"
_PROMPT_CATALOG_SCHEMA = "concurrent-robustness-v2-prompt-catalog-v1"
_CANDIDATE_TYPE = "prompt_model_realized_table_first_candidate"
_REPORT_HTML = "report.html"
_MANIFEST = "artifact_manifest.json"
_REPORT_PAYLOAD = "robustness_v2_report_payload.json"
_PROJECTION_JSON = "robustness_v2_projection.json"
_REALIZED_CSV = "robustness_v2_realized_main.csv"
_JUDGMENT_CSV = "robustness_v2_judgment_audit.csv"
_PROMPT_CATALOG_JSON = "robustness_v2_prompt_catalog.json"
_PROVIDER_CSV = "robustness_v2_provider_audit.csv"
_CELL_BATCH_CSV = "robustness_v2_cell_batch_evidence.csv"
_WORKBOOK = "robustness_v2_teacher_results.xlsx"
_MECHANISM_MMD = "prompt-model-realized-mechanism.mmd"
_SEGMENTS = ("S1", "S2", "S3")
_MESSAGE_LABELS = {"message_1": "M1", "message_2": "M2", "message_3": "M3"}
_PROMPTS = ("P0", "P1", "P2", "P3")
_POSITIVE_ACTIONS = ("like", "comment", "share")
_MAX_HTML_BYTES = 3 * 1024 * 1024
_FIXED_WORKBOOK_TIME = datetime(2000, 1, 1, 0, 0, 0)
_FIXED_ZIP_TIME = (1980, 1, 1, 0, 0, 0)

_REALIZED_FIELDS = (
    "model",
    "prompt",
    "segment",
    "message",
    "like_count",
    "comment_count",
    "share_count",
    "engagement_count",
    "exposure_count",
    "engagement_rate",
    "prompt_anchor",
)
_REALIZED_HEADERS = (
    "Model",
    "Prompt",
    "Segment",
    "Message",
    "Like count",
    "Comment count",
    "Share count",
    "Engagement count",
    "Exposure count",
    "Engagement rate",
    "Prompt anchor",
)
_JUDGMENT_FIELDS = (
    "model",
    "prompt",
    "segment",
    "message",
    "provider_like_count",
    "provider_comment_count",
    "provider_share_count",
    "provider_ignore_count",
    "positive_judgment_count",
    "logical_judgment_count",
    "positive_judgment_rate",
    "mean_probability",
    "mean_confidence",
    "terminal_failure_count",
    "physical_attempt_count",
    "retry_attempt_count",
    "provider_response_count",
    "usage_complete_judgment_count",
    "usage_missing_judgment_count",
    "input_tokens",
    "output_tokens",
    "total_tokens",
    "cached_input_tokens",
    "requested_model_counts",
    "observed_model_counts",
    "provider_route_counts",
    "billing_semantics_counts",
    "provider_fee_cny",
    "subscription_nominal_cost_usd_reference",
    "prompt_anchor",
)
_JUDGMENT_HEADERS = tuple(field.replace("_", " ").title() for field in _JUDGMENT_FIELDS)
_PROVIDER_FIELDS = (
    "execution_profile",
    "condition_evidence_scope",
    "requested_model",
    "required_observed_model",
    "observed_model_counts",
    "planned_provider_route",
    "observed_provider_route_counts",
    "planned_route_kind",
    "planned_wire_model",
    "planned_wire_api",
    "planned_reasoning_effort",
    "planned_thinking_mode",
    "planned_output_token_ceiling",
    "planned_billing_semantics",
    "observed_billing_semantics_counts",
    "planned_billing_currency",
    "planned_fee_ceiling",
    "persisted_request_contract",
    "physical_attempt_count",
    "terminal_failure_count",
    "input_tokens",
    "output_tokens",
    "total_tokens",
    "gateway_context_visibility",
    "direct_gemini_developer_api",
    "client_prompt_scope",
)
_PROVIDER_HEADERS = tuple(field.replace("_", " ").title() for field in _PROVIDER_FIELDS)
_CELL_BATCH_FIELDS = (
    "cell_index",
    "cell_id",
    "model",
    "prompt",
    "batch",
    "frozen_feedback_user_count",
    "M1_selected_count",
    "M1_realized_positive_count",
    "M2_selected_count",
    "M2_realized_positive_count",
    "M3_selected_count",
    "M3_realized_positive_count",
    "selected_pair_count",
    "realized_positive_pair_count",
    "committed_positive_user_count",
)
_CELL_BATCH_HEADERS = tuple(field.replace("_", " ").title() for field in _CELL_BATCH_FIELDS)
_PROMPT_FIELDS = (
    "variant_id",
    "controlled_change",
    "prompt_version",
    "canonical_hash",
    "system_template",
    "user_template",
    "decision_json_schema",
    "request_settings",
    "message_catalog",
)
_PROMPT_HEADERS = (
    "Prompt",
    "Controlled change",
    "Version",
    "SHA-256",
    "System template",
    "User template",
    "Decision JSON schema",
    "Request settings",
    "Three messages",
)


class ConcurrentRobustnessV2ReportError(ValueError):
    """A v2 table-first report candidate failed closed."""


@dataclass(frozen=True)
class _ValidatedReportProjection:
    source_lineage: Mapping[str, Any]
    formal_topology: Mapping[str, int]
    realized_denominator: Mapping[str, int]
    realized_main_rows: tuple[Mapping[str, Any], ...]
    judgment_audit_rows: tuple[Mapping[str, Any], ...]
    prompt_catalog: tuple[Mapping[str, Any], ...]
    message_catalog: tuple[Mapping[str, Any], ...]
    provider_audit_rows: tuple[Mapping[str, Any], ...]
    cell_batch_evidence_rows: tuple[Mapping[str, Any], ...]
    claim_boundary: Mapping[str, Any]
    mechanism_schema_version: str
    mechanism_identity_sha256: str

    def document(self) -> dict[str, Any]:
        return {
            "schema_version": _PROJECTION_SCHEMA,
            "primary_outcome": "abm_realized_engagement",
            "source_lineage": dict(self.source_lineage),
            "formal_topology": dict(self.formal_topology),
            "realized_denominator": dict(self.realized_denominator),
            "realized_main_rows": [dict(row) for row in self.realized_main_rows],
            "judgment_audit_rows": [dict(row) for row in self.judgment_audit_rows],
            "prompt_catalog": [dict(row) for row in self.prompt_catalog],
            "message_catalog": [dict(row) for row in self.message_catalog],
            "provider_audit_rows": [dict(row) for row in self.provider_audit_rows],
            "cell_batch_evidence_rows": [dict(row) for row in self.cell_batch_evidence_rows],
            "claim_boundary": dict(self.claim_boundary),
            "mechanism_presentation": {
                "schema_version": self.mechanism_schema_version,
                "semantic_set_identity_sha256": self.mechanism_identity_sha256,
                "mermaid_download": _MECHANISM_MMD,
            },
            "provider_calls_during_composition": 0,
            "canonical_deployment_triggered": False,
            "production_deploy_eligible": False,
        }


@dataclass(frozen=True)
class _V2RealizedCandidateFacts:
    root: Path
    manifest_sha256: str
    candidate_identity_sha256: str
    projection_sha256: str
    report_sha256: str
    workbook_sha256: str
    artifact_hashes: Mapping[str, str]
    approved_downloads: Mapping[str, str]


@dataclass(frozen=True)
class _ExpectedCandidate:
    new_payloads: Mapping[str, bytes]
    final_report: bytes
    manifest: Mapping[str, Any]
    approved_downloads: Mapping[str, str]
    artifact_hashes: Mapping[str, str]
    artifact_sizes: Mapping[str, int]


def _canonical_json_bytes(value: object) -> bytes:
    return (
        json.dumps(
            value,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
            allow_nan=False,
        )
        + "\n"
    ).encode("utf-8")


def _compact_json(value: object) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"), allow_nan=False)


def _sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _valid_sha256(value: object) -> bool:
    return isinstance(value, str) and re.fullmatch(r"[0-9a-f]{64}", value) is not None


def _strict_int(value: object, label: str) -> int:
    if type(value) is not int or value < 0:
        raise ConcurrentRobustnessV2ReportError(f"{label} must be a non-negative strict integer")
    return value


def _finite_float(value: object, label: str) -> float:
    if isinstance(value, bool) or not isinstance(value, (int, float)) or not math.isfinite(float(value)):
        raise ConcurrentRobustnessV2ReportError(f"{label} must be finite")
    return float(value)


def _mapping(value: object, label: str) -> dict[str, Any]:
    if not isinstance(value, Mapping):
        raise ConcurrentRobustnessV2ReportError(f"{label} must be an object")
    return {str(key): item for key, item in value.items()}


def _object_rows(value: object, label: str) -> list[dict[str, Any]]:
    if not isinstance(value, Sequence) or isinstance(value, (str, bytes, bytearray)):
        raise ConcurrentRobustnessV2ReportError(f"{label} must be a list")
    rows: list[dict[str, Any]] = []
    for item in value:
        rows.append(_mapping(item, label))
    return rows


def _tree_records(root: Path) -> tuple[dict[str, str], dict[str, int]]:
    if root.is_symlink() or not root.is_dir():
        raise ConcurrentRobustnessV2ReportError("candidate lineage must be a real directory")
    hashes: dict[str, str] = {}
    sizes: dict[str, int] = {}
    for path in sorted(root.rglob("*")):
        if path.is_symlink():
            raise ConcurrentRobustnessV2ReportError("candidate inventory contains a symlink")
        if path.is_dir():
            continue
        if not path.is_file():
            raise ConcurrentRobustnessV2ReportError("candidate inventory contains a non-regular artifact")
        relative = path.relative_to(root).as_posix()
        pure = PurePosixPath(relative)
        if pure.is_absolute() or ".." in pure.parts or pure.as_posix() != relative:
            raise ConcurrentRobustnessV2ReportError("candidate inventory contains an unsafe path")
        hashes[relative] = _sha256_file(path)
        sizes[relative] = path.stat().st_size
    return hashes, sizes


def _string_counter(value: object, label: str) -> dict[str, int]:
    mapping = _mapping(value, label)
    result: dict[str, int] = {}
    for key, item in mapping.items():
        if not key:
            raise ConcurrentRobustnessV2ReportError(f"{label} contains an empty key")
        result[key] = _strict_int(item, label)
    return dict(sorted(result.items()))


def _ordered_models(source: _ConcurrentRobustnessV2ReportSource) -> tuple[str, ...]:
    observed: list[str] = []
    for cell in source.manifest.prompt_model_cells:
        if cell.requested_model not in observed:
            observed.append(cell.requested_model)
    if tuple(observed) != _V2_MODELS:
        raise ConcurrentRobustnessV2ReportError("v2 report model ordering is crossed")
    return tuple(observed)


def _normalize_realized_rows(source: _ConcurrentRobustnessV2ReportSource) -> tuple[Mapping[str, Any], ...]:
    analysis = _mapping(source.realized_analysis, "v2 Realized analysis")
    rows = [
        row
        for row in _object_rows(analysis.get("group_rows"), "v2 Realized group rows")
        if row.get("scope") == "model_prompt_segment_message"
    ]
    models = _ordered_models(source)
    expected_keys = [
        (model, prompt, segment, message_id)
        for model in models
        for prompt in _PROMPTS
        for segment in _SEGMENTS
        for message_id in source.manifest.message_ids
    ]
    actual_keys = [
        (row.get("requested_model"), row.get("prompt_variant"), row.get("segment"), row.get("message_id"))
        for row in rows
    ]
    if actual_keys != expected_keys:
        raise ConcurrentRobustnessV2ReportError("Realized main row ordering or denominator is crossed")
    normalized: list[Mapping[str, Any]] = []
    for row in rows:
        like_count = _strict_int(row.get("like_count"), "Realized likes")
        comment_count = _strict_int(row.get("comment_count"), "Realized comments")
        share_count = _strict_int(row.get("share_count"), "Realized shares")
        engagement_count = _strict_int(row.get("engagement_count"), "Realized engagements")
        exposure_count = _strict_int(row.get("exposure_count"), "Realized exposures")
        rate = _finite_float(row.get("engagement_rate"), "Realized engagement rate")
        if (
            engagement_count != like_count + comment_count + share_count
            or not 0.0 <= rate <= 1.0
            or rate != (round(engagement_count / exposure_count, 12) if exposure_count else 0.0)
        ):
            raise ConcurrentRobustnessV2ReportError("Realized action identity or rate is crossed")
        prompt = str(row["prompt_variant"])
        normalized.append(
            {
                "model": str(row["requested_model"]),
                "prompt": prompt,
                "segment": str(row["segment"]),
                "message": str(row["message_label"]),
                "like_count": like_count,
                "comment_count": comment_count,
                "share_count": share_count,
                "engagement_count": engagement_count,
                "exposure_count": exposure_count,
                "engagement_rate": rate,
                "prompt_anchor": f"#prompt-catalog-{prompt}",
            }
        )
    return tuple(normalized)


def _normalize_judgment_rows(source: _ConcurrentRobustnessV2ReportSource) -> tuple[Mapping[str, Any], ...]:
    audit = _mapping(source.judgment_audit, "v2 Judgment audit")
    rows = [
        row
        for row in _object_rows(audit.get("group_rows"), "v2 Judgment group rows")
        if row.get("scope") == "model_prompt_segment_message"
    ]
    expected_keys = [
        (model, prompt, segment, message_id)
        for model in _ordered_models(source)
        for prompt in _PROMPTS
        for segment in _SEGMENTS
        for message_id in source.manifest.message_ids
    ]
    actual_keys = [
        (row.get("requested_model"), row.get("prompt_variant"), row.get("segment"), row.get("message_id"))
        for row in rows
    ]
    if actual_keys != expected_keys:
        raise ConcurrentRobustnessV2ReportError("Judgment audit row ordering or denominator is crossed")
    integer_fields = (
        "provider_like_count",
        "provider_comment_count",
        "provider_share_count",
        "provider_ignore_count",
        "positive_judgment_count",
        "logical_judgment_count",
        "terminal_failure_count",
        "physical_attempt_count",
        "retry_attempt_count",
        "provider_response_count",
        "usage_complete_judgment_count",
        "usage_missing_judgment_count",
        "input_tokens",
        "output_tokens",
        "total_tokens",
        "cached_input_tokens",
    )
    float_fields = (
        "positive_judgment_rate",
        "mean_probability",
        "mean_confidence",
        "provider_fee_cny",
        "subscription_nominal_cost_usd_reference",
    )
    counter_fields = (
        "requested_model_counts",
        "observed_model_counts",
        "provider_route_counts",
        "billing_semantics_counts",
    )
    normalized: list[Mapping[str, Any]] = []
    for row in rows:
        values: dict[str, Any] = {
            "model": str(row["requested_model"]),
            "prompt": str(row["prompt_variant"]),
            "segment": str(row["segment"]),
            "message": str(row["message_label"]),
        }
        values.update({field: _strict_int(row.get(field), f"Judgment {field}") for field in integer_fields})
        values.update({field: _finite_float(row.get(field), f"Judgment {field}") for field in float_fields})
        values.update(
            {
                field: _compact_json(_string_counter(row.get(field), f"Judgment {field}"))
                for field in counter_fields
            }
        )
        positive = sum(values[f"provider_{action}_count"] for action in _POSITIVE_ACTIONS)
        denominator = values["logical_judgment_count"]
        if (
            positive != values["positive_judgment_count"]
            or positive + values["provider_ignore_count"] != denominator
            or values["positive_judgment_rate"]
            != (round(positive / denominator, 12) if denominator else 0.0)
            or values["terminal_failure_count"] != 0
        ):
            raise ConcurrentRobustnessV2ReportError("Judgment action identity or denominator is crossed")
        values["prompt_anchor"] = f"#prompt-catalog-{values['prompt']}"
        normalized.append(values)
    return tuple(normalized)


def _message_catalog(source: _ConcurrentRobustnessV2ReportSource) -> tuple[Mapping[str, Any], ...]:
    rows = [_mapping(row, "v2 message snapshot row") for row in source.message_snapshot]
    if tuple(row.get("message_id") for row in rows) != source.manifest.message_ids:
        raise ConcurrentRobustnessV2ReportError("v2 message snapshot order is crossed")
    result: list[Mapping[str, Any]] = []
    for row in rows:
        message_id = str(row["message_id"])
        title = row.get("title")
        body = row.get("body")
        segment = row.get("intended_audience_segment")
        dimensions = row.get("value_dimensions")
        if (
            not isinstance(title, str)
            or not title.strip()
            or not isinstance(body, str)
            or not body.strip()
            or segment not in {"class_1", "class_2", "class_3"}
            or not isinstance(dimensions, Mapping)
        ):
            raise ConcurrentRobustnessV2ReportError("v2 message catalog contains malformed content")
        result.append(
            {
                "message_id": message_id,
                "message": _MESSAGE_LABELS[message_id],
                "title": title,
                "intended_audience_segment": segment,
                "body": body,
                "value_dimensions": dict(dimensions),
            }
        )
    return tuple(result)


def _provider_audit_rows(source: _ConcurrentRobustnessV2ReportSource) -> tuple[Mapping[str, Any], ...]:
    audit = _mapping(source.judgment_audit, "v2 Judgment audit")
    model_rows = {
        str(row["requested_model"]): row
        for row in _object_rows(audit.get("group_rows"), "v2 Judgment group rows")
        if row.get("scope") == "model" and isinstance(row.get("requested_model"), str)
    }
    requirements: dict[str, str] = {}
    for cell in source.manifest.prompt_model_cells:
        required = cell.required_observed_model
        if required is None:
            raise ConcurrentRobustnessV2ReportError("Provider condition is missing required observed identity")
        previous = requirements.setdefault(cell.requested_model, required)
        if previous != required:
            raise ConcurrentRobustnessV2ReportError("Provider observed identity differs across Prompt cells")
    disclosures = robustness_provider_disclosures()
    if tuple(row["requested_model"] for row in disclosures) != _ordered_models(source):
        raise ConcurrentRobustnessV2ReportError("Provider disclosure ordering is crossed")
    persisted_request_contract = source.manifest.request_contract.model_dump(mode="json")
    result: list[Mapping[str, Any]] = []
    for disclosure in disclosures:
        model = str(disclosure["requested_model"])
        audit_row = model_rows.get(model)
        if audit_row is None:
            raise ConcurrentRobustnessV2ReportError("Provider audit is missing a model row")
        required_observed_model = requirements.get(model)
        if required_observed_model != disclosure["required_observed_model"]:
            raise ConcurrentRobustnessV2ReportError("planned Provider identity differs from persisted evidence")
        physical_attempts = _strict_int(
            audit_row.get("physical_attempt_count"), "Provider physical attempts"
        )
        logical_judgments = _strict_int(
            audit_row.get("logical_judgment_count"), "Provider logical judgments"
        )
        observed_models = _string_counter(
            audit_row.get("observed_model_counts"), "observed model counts"
        )
        observed_routes = _string_counter(
            audit_row.get("provider_route_counts"), "Provider route counts"
        )
        observed_billing = _string_counter(
            audit_row.get("billing_semantics_counts"), "Provider billing semantics counts"
        )
        planned_route = str(disclosure["provider_route"])
        planned_billing = str(disclosure["billing_semantics"])
        planned_condition_observed = (
            observed_routes == {planned_route: physical_attempts}
            and observed_billing == {planned_billing: physical_attempts}
        )
        deterministic_validation_observed = (
            observed_routes == {"injected_deterministic_validation": physical_attempts}
            and observed_billing == {"none": physical_attempts}
        )
        if source.manifest.execution_profile == "formal":
            if not planned_condition_observed:
                raise ConcurrentRobustnessV2ReportError(
                    "Formal Provider route or billing evidence differs from its planned condition"
                )
            condition_scope = "planned_condition_observed_in_closed_formal_evidence"
        elif deterministic_validation_observed:
            condition_scope = "planned_condition_not_executed_in_deterministic_validation"
        elif planned_condition_observed:
            condition_scope = "planned_condition_observed_in_offline_validation"
        else:
            raise ConcurrentRobustnessV2ReportError(
                "validation Provider evidence matches neither its injected nor planned condition"
            )
        if observed_models != {required_observed_model: logical_judgments}:
            raise ConcurrentRobustnessV2ReportError("observed model evidence differs from the required identity")
        result.append(
            {
                "execution_profile": source.manifest.execution_profile,
                "condition_evidence_scope": condition_scope,
                "requested_model": model,
                "required_observed_model": required_observed_model,
                "observed_model_counts": _compact_json(observed_models),
                "planned_provider_route": planned_route,
                "observed_provider_route_counts": _compact_json(observed_routes),
                "planned_route_kind": str(disclosure["route_kind"]),
                "planned_wire_model": str(disclosure["wire_model"]),
                "planned_wire_api": str(disclosure["wire_api"]),
                "planned_reasoning_effort": disclosure["reasoning_effort"],
                "planned_thinking_mode": disclosure["thinking_mode"],
                "planned_output_token_ceiling": _strict_int(
                    disclosure["output_token_ceiling"], "Provider output-token ceiling"
                ),
                "planned_billing_semantics": planned_billing,
                "observed_billing_semantics_counts": _compact_json(observed_billing),
                "planned_billing_currency": disclosure["billing_currency"],
                "planned_fee_ceiling": disclosure["fee_ceiling"],
                "persisted_request_contract": _compact_json(persisted_request_contract),
                "physical_attempt_count": physical_attempts,
                "terminal_failure_count": _strict_int(
                    audit_row.get("terminal_failure_count"), "Provider failures"
                ),
                "input_tokens": _strict_int(audit_row.get("input_tokens"), "Provider input tokens"),
                "output_tokens": _strict_int(audit_row.get("output_tokens"), "Provider output tokens"),
                "total_tokens": _strict_int(audit_row.get("total_tokens"), "Provider total tokens"),
                "gateway_context_visibility": str(disclosure["gateway_context_visibility"]),
                "direct_gemini_developer_api": disclosure["direct_gemini_developer_api"],
                "client_prompt_scope": str(disclosure["client_prompt_scope"]),
            }
        )
    return tuple(result)


def _prompt_catalog(
    provider_rows: Sequence[Mapping[str, Any]],
    messages: Sequence[Mapping[str, Any]],
) -> tuple[Mapping[str, Any], ...]:
    request_setting_fields = (
        "requested_model",
        "planned_provider_route",
        "planned_route_kind",
        "planned_wire_model",
        "planned_wire_api",
        "planned_reasoning_effort",
        "planned_thinking_mode",
        "planned_output_token_ceiling",
        "persisted_request_contract",
        "gateway_context_visibility",
        "direct_gemini_developer_api",
        "client_prompt_scope",
    )
    request_settings = [
        {field: row[field] for field in request_setting_fields}
        for row in provider_rows
    ]
    message_records = [dict(row) for row in messages]
    result: list[Mapping[str, Any]] = []
    for record in CONCURRENT_ROBUSTNESS_PROMPT_REGISTRY.catalog_records():
        templates = cast(tuple[dict[str, str], ...], record["client_submitted_message_templates"])
        if tuple(item["role"] for item in templates) != ("system", "user"):
            raise ConcurrentRobustnessV2ReportError("Prompt catalog message roles are crossed")
        result.append(
            {
                **record,
                "client_submitted_message_templates": [dict(item) for item in templates],
                "request_settings": request_settings,
                "messages": message_records,
            }
        )
    return tuple(result)


def _cell_batch_rows(source: _ConcurrentRobustnessV2ReportSource) -> tuple[Mapping[str, Any], ...]:
    result: list[Mapping[str, Any]] = []
    expected_keys = [
        (cell_index, time_step)
        for cell_index in range(len(source.manifest.prompt_model_cells))
        for time_step in range(source.manifest.ranking_contract.horizon)
    ]
    actual_keys: list[tuple[int, int]] = []
    for raw in source.batch_commits:
        row = _mapping(raw, "v2 batch commit")
        cell_index = _strict_int(row.get("cell_index"), "cell index")
        time_step = _strict_int(row.get("time_step"), "batch time step")
        actual_keys.append((cell_index, time_step))
        cell = source.manifest.prompt_model_cells[cell_index]
        messages = _object_rows(row.get("messages"), "v2 batch messages")
        if tuple(item.get("message_id") for item in messages) != source.manifest.message_ids:
            raise ConcurrentRobustnessV2ReportError("Cell-batch message order is crossed")
        normalized: dict[str, Any] = {
            "cell_index": cell_index,
            "cell_id": cell.cell_id,
            "model": cell.requested_model,
            "prompt": cell.prompt_variant,
            "batch": time_step + 1,
            "frozen_feedback_user_count": len(
                cast(Sequence[object], row["frozen_campaign_engaged_user_ids"])
            ),
        }
        selected_total = 0
        positive_total = 0
        for message in messages:
            message_id = str(message["message_id"])
            label = _MESSAGE_LABELS[message_id]
            selected = cast(Sequence[object], message["selected_user_ids"])
            positives = cast(Sequence[object], message["realized_positive_user_ids"])
            normalized[f"{label}_selected_count"] = len(selected)
            normalized[f"{label}_realized_positive_count"] = len(positives)
            selected_total += len(selected)
            positive_total += len(positives)
        normalized["selected_pair_count"] = selected_total
        normalized["realized_positive_pair_count"] = positive_total
        normalized["committed_positive_user_count"] = len(
            cast(Sequence[object], row["committed_realized_positive_user_ids"])
        )
        result.append(normalized)
    if actual_keys != expected_keys:
        raise ConcurrentRobustnessV2ReportError("Cell-batch evidence is missing or reordered")
    return tuple(result)


def _build_projection(
    *,
    source: _ConcurrentRobustnessV2ReportSource,
    source_lineage: Mapping[str, Any],
) -> _ValidatedReportProjection:
    analysis = _mapping(source.realized_analysis, "v2 Realized analysis")
    audit = _mapping(source.judgment_audit, "v2 Judgment audit")
    claims = _mapping(source.claim_audit, "v2 claim audit")
    formal_topology = _mapping(analysis.get("formal_topology"), "v2 Formal topology")
    realized_denominator = _mapping(analysis.get("realized_denominator"), "v2 Realized denominator")
    expected_formal = {
        "cells": 20,
        "logical_judgments_per_cell": 1_800,
        "logical_judgments": 36_000,
        "maximum_physical_attempts": 108_000,
    }
    actual_denominator = {
        "cells": _strict_int(realized_denominator.get("cells"), "Realized cell count"),
        "logical_judgments_per_cell": _strict_int(
            realized_denominator.get("logical_judgments_per_cell"),
            "Realized judgments per cell",
        ),
        "logical_judgments": _strict_int(
            realized_denominator.get("logical_judgments"), "Realized judgments"
        ),
        "exposures": _strict_int(realized_denominator.get("exposures"), "Realized exposures"),
    }
    if (
        formal_topology != expected_formal
        or actual_denominator["cells"] != 20
        or actual_denominator["logical_judgments"] != source.facts.logical_judgments
        or actual_denominator["exposures"] != source.facts.logical_judgments
        or audit.get("scope") != "provider_judgment_only"
        or claims.get("status") != "passed"
        or source.facts.provider_calls != audit.get("counts", {}).get("provider_calls")
    ):
        raise ConcurrentRobustnessV2ReportError("v2 report source denominators or evidence scopes are crossed")
    realized_rows = _normalize_realized_rows(source)
    judgment_rows = _normalize_judgment_rows(source)
    messages = _message_catalog(source)
    providers = _provider_audit_rows(source)
    prompts = _prompt_catalog(providers, messages)
    cell_batches = _cell_batch_rows(source)
    if (
        sum(int(row["exposure_count"]) for row in realized_rows) != source.facts.logical_judgments
        or sum(int(row["logical_judgment_count"]) for row in judgment_rows)
        != source.facts.logical_judgments
    ):
        raise ConcurrentRobustnessV2ReportError("table projection does not close its Realized or Judgment denominator")
    mechanism = _MECHANISM_PRESENTATION.build_robustness_v2_master()
    claim_boundary = {
        "fixed_sample": True,
        "fixed_graph": True,
        "shared_deterministic_draw": True,
        "one_realized_path_per_cell": True,
        "historical_sixteen_cells_scope": "immutable_judgment_reference_only",
        "historical_cells_in_realized_main": False,
        "provider_reason_scope": "judgment_only",
        "invented_realized_explanation": False,
        "winner_claim": False,
        "accuracy_or_calibration_claim": False,
        "causal_or_external_validity_claim": False,
    }
    projection = _ValidatedReportProjection(
        source_lineage=dict(source_lineage),
        formal_topology=cast(Mapping[str, int], expected_formal),
        realized_denominator=actual_denominator,
        realized_main_rows=realized_rows,
        judgment_audit_rows=judgment_rows,
        prompt_catalog=prompts,
        message_catalog=messages,
        provider_audit_rows=providers,
        cell_batch_evidence_rows=cell_batches,
        claim_boundary=claim_boundary,
        mechanism_schema_version=mechanism.schema_version,
        mechanism_identity_sha256=mechanism.semantic_set_identity_sha256,
    )
    serialized = _canonical_json_bytes(projection.document())
    if b"realized_reason" in serialized or b"rendered_prompt" in serialized:
        raise ConcurrentRobustnessV2ReportError("report projection contains a forbidden duplicated evidence field")
    return projection


def _csv_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (dict, list, tuple)):
        return _compact_json(value)
    if isinstance(value, float):
        if not math.isfinite(value):
            raise ConcurrentRobustnessV2ReportError("CSV contains a non-finite number")
        return format(value, ".12g")
    return value


def _csv_bytes(fields: Sequence[str], rows: Sequence[Mapping[str, Any]]) -> bytes:
    stream = io.StringIO(newline="")
    writer = csv.DictWriter(stream, fieldnames=list(fields), lineterminator="\n", extrasaction="raise")
    writer.writeheader()
    for row in rows:
        writer.writerow({field: _csv_value(row.get(field)) for field in fields})
    return stream.getvalue().encode("utf-8")


def _prompt_sheet_rows(projection: _ValidatedReportProjection) -> tuple[tuple[object, ...], ...]:
    rows: list[tuple[object, ...]] = []
    for record in projection.prompt_catalog:
        messages = cast(Sequence[Mapping[str, str]], record["client_submitted_message_templates"])
        rows.append(
            (
                record["variant_id"],
                record["controlled_change"],
                record["prompt_version"],
                record["canonical_hash"],
                messages[0]["content"],
                messages[1]["content"],
                _compact_json(record["decision_output_schema"]),
                _compact_json(record["request_settings"]),
                _compact_json(record["messages"]),
            )
        )
    return tuple(rows)


def _readme_rows(projection: _ValidatedReportProjection) -> tuple[tuple[object, ...], ...]:
    return (
        ("Schema", _PROJECTION_SCHEMA),
        ("Primary result", "ABM Realized like/comment/share/engagement/exposure and engagement rate"),
        ("Judgment scope", "Provider Judgment remains in the separate Judgment Audit sheet"),
        ("Grouping", "Model → Prompt → S1–S3 → M1–M3"),
        ("Formal topology", _compact_json(projection.formal_topology)),
        ("Validated denominator", _compact_json(projection.realized_denominator)),
        ("Source lineage", _compact_json(projection.source_lineage)),
        ("Claim boundary", _compact_json(projection.claim_boundary)),
        ("Workbook contract", "deterministic-xlsxwriter-with-independent-openpyxl-reread-v1"),
        ("Production deploy eligible", "false"),
        ("Provider calls during composition", 0),
    )


def _workbook_tables(
    projection: _ValidatedReportProjection,
) -> tuple[tuple[str, tuple[str, ...], tuple[tuple[object, ...], ...]], ...]:
    realized_rows = tuple(tuple(row[field] for field in _REALIZED_FIELDS) for row in projection.realized_main_rows)
    judgment_rows = tuple(tuple(row[field] for field in _JUDGMENT_FIELDS) for row in projection.judgment_audit_rows)
    provider_rows = tuple(tuple(row[field] for field in _PROVIDER_FIELDS) for row in projection.provider_audit_rows)
    cell_batch_rows = tuple(
        tuple(row[field] for field in _CELL_BATCH_FIELDS) for row in projection.cell_batch_evidence_rows
    )
    return (
        ("README & Lineage", ("Field", "Value"), _readme_rows(projection)),
        ("Realized Main", _REALIZED_HEADERS, realized_rows),
        ("Judgment Audit", _JUDGMENT_HEADERS, judgment_rows),
        ("Prompt Catalog", _PROMPT_HEADERS, _prompt_sheet_rows(projection)),
        ("Provider Audit", _PROVIDER_HEADERS, provider_rows),
        ("Cell-Batch Evidence", _CELL_BATCH_HEADERS, cell_batch_rows),
    )


def _xlsx_cell_value(value: object) -> object:
    if isinstance(value, (dict, list, tuple)):
        return _compact_json(value)
    return value


def _normalize_xlsx_zip(payload: bytes) -> bytes:
    source = io.BytesIO(payload)
    destination = io.BytesIO()
    with zipfile.ZipFile(source, "r") as archive, zipfile.ZipFile(
        destination,
        "w",
        compression=zipfile.ZIP_DEFLATED,
        compresslevel=9,
    ) as output:
        for name in sorted(archive.namelist()):
            info = zipfile.ZipInfo(name, date_time=_FIXED_ZIP_TIME)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 3
            info.external_attr = 0o600 << 16
            output.writestr(info, archive.read(name))
    return destination.getvalue()


def _workbook_bytes(projection: _ValidatedReportProjection) -> bytes:
    stream = io.BytesIO()
    workbook = xlsxwriter.Workbook(
        stream,
        {
            "in_memory": True,
            "strings_to_formulas": False,
            "strings_to_urls": False,
        },
    )
    workbook.set_properties(
        {
            "title": "Prompt–Model Realized Robustness Teacher Results",
            "subject": "Manifest-bound table-first bilingual research results",
            "author": "llm-abm-marketing-sim Report Module",
            "company": "llm-abm-marketing-sim",
            "comments": "Deterministic workbook; Provider Judgment and ABM Realized facts remain separate.",
            "created": _FIXED_WORKBOOK_TIME,
        }
    )
    header_format = workbook.add_format(
        {
            "bold": True,
            "font_color": "#FFFFFF",
            "bg_color": "#163456",
            "border": 1,
            "align": "center",
            "valign": "vcenter",
        }
    )
    text_format = workbook.add_format({"valign": "top"})
    wrapped_format = workbook.add_format({"text_wrap": True, "valign": "top"})
    rate_format = workbook.add_format({"num_format": "0.0000%", "valign": "top"})
    for sheet_name, headers, rows in _workbook_tables(projection):
        worksheet = workbook.add_worksheet(sheet_name)
        worksheet.freeze_panes(1, 0)
        worksheet.set_row(0, 24)
        for column, heading in enumerate(headers):
            worksheet.write(0, column, heading, header_format)
        for row_index, row in enumerate(rows, start=1):
            for column, raw in enumerate(row):
                value = _xlsx_cell_value(raw)
                cell_format = text_format
                if sheet_name == "Realized Main" and headers[column] == "Engagement rate":
                    cell_format = rate_format
                elif sheet_name in {"Judgment Audit", "Prompt Catalog"} and (
                    len(str(value)) > 120 or "Template" in headers[column] or "Settings" in headers[column]
                ):
                    cell_format = wrapped_format
                if value is None:
                    worksheet.write_blank(row_index, column, None, cell_format)
                elif isinstance(value, bool):
                    worksheet.write_boolean(row_index, column, value, cell_format)
                elif isinstance(value, int):
                    worksheet.write_number(row_index, column, value, cell_format)
                elif isinstance(value, float):
                    if not math.isfinite(value):
                        raise ConcurrentRobustnessV2ReportError("workbook contains a non-finite number")
                    worksheet.write_number(row_index, column, value, cell_format)
                else:
                    worksheet.write_string(row_index, column, str(value), cell_format)
        last_row = max(1, len(rows))
        worksheet.autofilter(0, 0, last_row, len(headers) - 1)
        if sheet_name == "Prompt Catalog":
            worksheet.set_column(0, 3, 24)
            worksheet.set_column(4, 8, 58)
            for row_index in range(1, len(rows) + 1):
                worksheet.set_row(row_index, 96)
        elif sheet_name == "README & Lineage":
            worksheet.set_column(0, 0, 30)
            worksheet.set_column(1, 1, 110, wrapped_format)
        else:
            worksheet.set_column(0, len(headers) - 1, 20)
    workbook.close()
    first = _normalize_xlsx_zip(stream.getvalue())
    with zipfile.ZipFile(io.BytesIO(first), "r") as archive:
        if any(info.date_time != _FIXED_ZIP_TIME for info in archive.infolist()):
            raise ConcurrentRobustnessV2ReportError("workbook ZIP metadata is not deterministic")
    return first


def _validate_workbook(payload: bytes, projection: _ValidatedReportProjection) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(payload), "r") as archive:
            if any(info.date_time != _FIXED_ZIP_TIME for info in archive.infolist()):
                raise ConcurrentRobustnessV2ReportError("workbook ZIP timestamps are crossed")
        workbook = load_workbook(io.BytesIO(payload), data_only=False, read_only=False)
    except ConcurrentRobustnessV2ReportError:
        raise
    except Exception as exc:
        raise ConcurrentRobustnessV2ReportError("workbook reader rejected the XLSX bytes") from exc
    try:
        tables = _workbook_tables(projection)
        if workbook.sheetnames != [table[0] for table in tables]:
            raise ConcurrentRobustnessV2ReportError("workbook sheet order is crossed")
        if (
            workbook.properties.title != "Prompt–Model Realized Robustness Teacher Results"
            or workbook.properties.creator != "llm-abm-marketing-sim Report Module"
            or workbook.properties.created != _FIXED_WORKBOOK_TIME
        ):
            raise ConcurrentRobustnessV2ReportError("workbook properties are crossed")
        for sheet_name, headers, rows in tables:
            worksheet = workbook[sheet_name]
            if worksheet.freeze_panes != "A2":
                raise ConcurrentRobustnessV2ReportError(f"workbook freeze pane is crossed: {sheet_name}")
            expected_filter = f"A1:{get_column_letter(len(headers))}{max(2, len(rows) + 1)}"
            if worksheet.auto_filter.ref != expected_filter:
                raise ConcurrentRobustnessV2ReportError(f"workbook autofilter is crossed: {sheet_name}")
            observed_headers = tuple(cell.value for cell in worksheet[1])
            if observed_headers != headers:
                raise ConcurrentRobustnessV2ReportError(f"workbook headers are crossed: {sheet_name}")
            observed_rows = tuple(
                tuple(cell.value for cell in row[: len(headers)])
                for row in worksheet.iter_rows(min_row=2, max_row=len(rows) + 1)
            )
            expected_rows = tuple(tuple(_xlsx_cell_value(value) for value in row) for row in rows)
            if observed_rows != expected_rows:
                raise ConcurrentRobustnessV2ReportError(f"workbook values are crossed: {sheet_name}")
            for observed_row_index, row in enumerate(
                worksheet.iter_rows(min_row=2, max_row=len(rows) + 1)
            ):
                for index, cell in enumerate(row[: len(headers)]):
                    expected = expected_rows[observed_row_index][index]
                    if isinstance(expected, int) and not isinstance(expected, bool) and type(cell.value) is not int:
                        raise ConcurrentRobustnessV2ReportError(
                            f"workbook integer type is crossed: {sheet_name}!{cell.coordinate}"
                        )
                    if isinstance(expected, float) and not isinstance(cell.value, (int, float)):
                        raise ConcurrentRobustnessV2ReportError(
                            f"workbook numeric type is crossed: {sheet_name}!{cell.coordinate}"
                        )
    finally:
        workbook.close()


def _display(value: object) -> str:
    if value is None:
        return "—"
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float):
        return format(value, ".12g")
    return str(value)


def _bilingual(zh_cn: str, en_us: str) -> str:
    return (
        f'<span data-v2-lang="zh-CN">{html.escape(zh_cn)}</span>'
        f'<span data-v2-lang="en-US" hidden>{html.escape(en_us)}</span>'
    )


def _table_header(zh_cn: str, en_us: str) -> str:
    return f"<th scope=\"col\">{_bilingual(zh_cn, en_us)}</th>"


def _realized_panels(projection: _ValidatedReportProjection) -> str:
    panels: list[str] = []
    for model in _V2_MODELS:
        for prompt in _PROMPTS:
            selected = [
                row for row in projection.realized_main_rows if row["model"] == model and row["prompt"] == prompt
            ]
            segments: list[str] = []
            for segment in _SEGMENTS:
                rows = [row for row in selected if row["segment"] == segment]
                body = "".join(
                    "<tr>"
                    f'<td>{html.escape(str(row["message"]))}</td>'
                    f'<td><a href="{html.escape(str(row["prompt_anchor"]), quote=True)}">{prompt}</a></td>'
                    f'<td>{row["like_count"]}</td><td>{row["comment_count"]}</td><td>{row["share_count"]}</td>'
                    f'<td>{row["engagement_count"]}</td><td>{row["exposure_count"]}</td>'
                    f'<td>{float(row["engagement_rate"]):.4%}</td></tr>'
                    for row in rows
                )
                segments.append(
                    f'<section class="robustness-v2-segment" data-v2-segment="{segment}"><h5>{segment}</h5>'
                    '<div class="robustness-v2-table-wrap"><table><thead><tr>'
                    + _table_header("Message", "Message")
                    + _table_header("Prompt", "Prompt")
                    + _table_header("点赞", "Like count")
                    + _table_header("评论", "Comment count")
                    + _table_header("分享", "Share count")
                    + _table_header("互动", "Engagement count")
                    + _table_header("曝光", "Exposure count")
                    + _table_header("互动率", "Engagement rate")
                    + f"</tr></thead><tbody>{body}</tbody></table></div></section>"
                )
            hidden = "" if model == _V2_MODELS[0] and prompt == _PROMPTS[0] else " hidden"
            panels.append(
                f'<article data-v2-result-panel data-v2-model="{html.escape(model, quote=True)}" '
                f'data-v2-prompt="{prompt}"{hidden}><h4>{html.escape(model)} · {prompt}</h4>{"".join(segments)}</article>'
            )
    return "".join(panels)


def _judgment_panels(projection: _ValidatedReportProjection) -> str:
    panels: list[str] = []
    for model in _V2_MODELS:
        for prompt in _PROMPTS:
            selected = [
                row for row in projection.judgment_audit_rows if row["model"] == model and row["prompt"] == prompt
            ]
            segments: list[str] = []
            for segment in _SEGMENTS:
                rows = [row for row in selected if row["segment"] == segment]
                body = "".join(
                    "<tr>"
                    f'<td>{html.escape(str(row["message"]))}</td>'
                    f'<td><a href="{html.escape(str(row["prompt_anchor"]), quote=True)}">{prompt}</a></td>'
                    f'<td>{row["provider_like_count"]}</td><td>{row["provider_comment_count"]}</td>'
                    f'<td>{row["provider_share_count"]}</td><td>{row["provider_ignore_count"]}</td>'
                    f'<td>{float(row["positive_judgment_rate"]):.4%}</td>'
                    f'<td>{float(row["mean_probability"]):.4f}</td><td>{float(row["mean_confidence"]):.4f}</td>'
                    f'<td>{row["terminal_failure_count"]}</td><td>{row["physical_attempt_count"]}</td>'
                    f'<td>{row["total_tokens"]}</td><td><code>{html.escape(str(row["observed_model_counts"]))}</code></td></tr>'
                    for row in rows
                )
                segments.append(
                    f'<section class="robustness-v2-segment" data-v2-segment="{segment}"><h5>{segment}</h5>'
                    '<div class="robustness-v2-table-wrap"><table><thead><tr>'
                    + _table_header("Message", "Message")
                    + _table_header("Prompt", "Prompt")
                    + _table_header("Judgment 点赞", "Judgment likes")
                    + _table_header("Judgment 评论", "Judgment comments")
                    + _table_header("Judgment 分享", "Judgment shares")
                    + _table_header("Judgment 忽略", "Judgment ignores")
                    + _table_header("正向率", "Positive rate")
                    + _table_header("平均概率", "Mean probability")
                    + _table_header("平均置信度", "Mean confidence")
                    + _table_header("失败", "Failures")
                    + _table_header("Attempts", "Attempts")
                    + _table_header("Usage", "Usage")
                    + _table_header("Observed model", "Observed model")
                    + f"</tr></thead><tbody>{body}</tbody></table></div></section>"
                )
            hidden = "" if model == _V2_MODELS[0] and prompt == _PROMPTS[0] else " hidden"
            panels.append(
                f'<article data-v2-judgment-panel data-v2-model="{html.escape(model, quote=True)}" '
                f'data-v2-prompt="{prompt}"{hidden}><h4>{html.escape(model)} · {prompt}</h4>{"".join(segments)}</article>'
            )
    return "".join(panels)


def _prompt_catalog_html(projection: _ValidatedReportProjection) -> str:
    records: list[str] = []
    for record in projection.prompt_catalog:
        messages = cast(Sequence[Mapping[str, str]], record["client_submitted_message_templates"])
        records.append(
            f'<article class="robustness-v2-prompt" id="prompt-catalog-{record["variant_id"]}" '
            f'data-testid="robustness-v2-prompt-{str(record["variant_id"]).lower()}">'
            f'<h4>{record["variant_id"]} · {html.escape(str(record["controlled_change"]))}</h4>'
            f'<p><code>{html.escape(str(record["prompt_version"]))}</code><br><code>{html.escape(str(record["canonical_hash"]))}</code></p>'
            f'<h5>System template</h5><pre>{html.escape(messages[0]["content"])}</pre>'
            f'<h5>User template</h5><pre>{html.escape(messages[1]["content"])}</pre>'
            f'<h5>Decision JSON schema</h5><pre>{html.escape(_compact_json(record["decision_output_schema"]))}</pre>'
            "</article>"
        )
    message_cards = "".join(
        f'<details><summary>{row["message"]} · {html.escape(str(row["title"]))}</summary>'
        f'<p><code>{html.escape(str(row["message_id"]))}</code> · {html.escape(str(row["intended_audience_segment"]))}</p>'
        f'<pre>{html.escape(str(row["body"]))}</pre><pre>{html.escape(_compact_json(row["value_dimensions"]))}</pre></details>'
        for row in projection.message_catalog
    )
    return (
        '<section class="robustness-v2-block" id="robustness-v2-prompt-catalog" '
        'data-testid="robustness-v2-prompt-catalog">'
        f'<h3>{_bilingual("Prompt Catalog：完整 client-submitted templates", "Prompt Catalog: complete client-submitted templates")}</h3>'
        f'<p>{_bilingual("这里展示静态 system/user templates、P0–P3 差异、版本与 hash；不保存逐用户 rendered Prompt。", "This catalog shows static system/user templates, P0–P3 differences, versions, and hashes; no per-user rendered Prompt is persisted.")}</p>'
        f'<div class="robustness-v2-prompt-grid">{"".join(records)}</div>'
        f'<h4>{_bilingual("三条 authoritative messages", "Three authoritative messages")}</h4>{message_cards}'
        "</section>"
    )


def _provider_audit_html(projection: _ValidatedReportProjection) -> str:
    rows: list[str] = []
    limitations: list[str] = []
    for row in projection.provider_audit_rows:
        rows.append(
            "<tr>"
            f'<td>{html.escape(str(row["execution_profile"]))}</td>'
            f'<td>{html.escape(str(row["condition_evidence_scope"]))}</td>'
            f'<td>{html.escape(str(row["requested_model"]))}</td>'
            f'<td>{html.escape(str(row["required_observed_model"]))}</td>'
            f'<td><code>{html.escape(str(row["observed_model_counts"]))}</code></td>'
            f'<td>{html.escape(str(row["planned_provider_route"]))}</td>'
            f'<td><code>{html.escape(str(row["observed_provider_route_counts"]))}</code></td>'
            f'<td>{html.escape(str(row["planned_reasoning_effort"] or row["planned_thinking_mode"] or "default"))}</td>'
            f'<td>{row["physical_attempt_count"]}</td><td>{row["total_tokens"]}</td>'
            f'<td>{html.escape(str(row["planned_billing_semantics"]))}</td>'
            f'<td><code>{html.escape(str(row["observed_billing_semantics_counts"]))}</code></td></tr>'
        )
        if row["gateway_context_visibility"] == "may_include_unobservable_context":
            limitations.append(
                f'<li><strong>{html.escape(str(row["requested_model"]))}</strong>: '
                'Antigravity OpenAI-compatible gateway; not direct Gemini Developer API. '
                'Gateway 可能注入不可见上下文；这里披露的是 client-submitted Prompt，'
                '不是完整 effective Prompt。</li>'
            )
    return (
        '<section class="robustness-v2-block" data-testid="robustness-v2-provider-audit">'
        f'<h3>{_bilingual("Provider Audit", "Provider Audit")}</h3>'
        f'<p>{_bilingual("Planned Provider condition 列来自冻结的 Provider contract；Observed 列只来自 closed execution evidence。deterministic_validation 不得伪装成 Formal route。", "Planned Provider condition columns come from the frozen Provider contract; Observed columns come only from closed execution evidence. Deterministic validation is never presented as a Formal route.")}</p>'
        '<div class="robustness-v2-table-wrap"><table><thead><tr>'
        + _table_header("Execution profile", "Execution profile")
        + _table_header("Condition evidence", "Condition evidence")
        + _table_header("Requested model", "Requested model")
        + _table_header("Required observed", "Required observed")
        + _table_header("Observed identity", "Observed identity")
        + _table_header("Planned route", "Planned route")
        + _table_header("Observed routes", "Observed routes")
        + _table_header("Planned reasoning / thinking", "Planned reasoning / thinking")
        + _table_header("Observed attempts", "Observed attempts")
        + _table_header("Observed usage", "Observed usage")
        + _table_header("Planned billing", "Planned billing")
        + _table_header("Observed billing", "Observed billing")
        + f'</tr></thead><tbody>{"".join(rows)}</tbody></table></div>'
        f'<aside class="robustness-v2-gateway-limit" data-testid="robustness-v2-gateway-limit"><ul>{"".join(limitations)}</ul></aside>'
        "</section>"
    )


def _download_mapping(base_report: str) -> dict[str, str]:
    existing_hrefs = sorted(
        {
            html.unescape(match)
            for match in re.findall(
                r'<a\b(?=[^>]*\bdownload\b)[^>]*\bhref="([^"]+)"',
                base_report,
                re.IGNORECASE,
            )
        }
    )
    result = {
        "teacher_results_xlsx": _WORKBOOK,
        "validated_projection_json": _PROJECTION_JSON,
        "realized_main_csv": _REALIZED_CSV,
        "judgment_audit_csv": _JUDGMENT_CSV,
        "prompt_catalog_json": _PROMPT_CATALOG_JSON,
        "provider_audit_csv": _PROVIDER_CSV,
        "cell_batch_evidence_csv": _CELL_BATCH_CSV,
        "prompt_model_realized_mechanism_mermaid": _MECHANISM_MMD,
        "report_payload_json": _REPORT_PAYLOAD,
    }
    known = set(result.values())
    for index, relative in enumerate(path for path in existing_hrefs if path not in known):
        result[f"closed_lineage_download_{index:03d}"] = relative
    if len(result) != len(set(result.values())):
        raise ConcurrentRobustnessV2ReportError("report download paths are duplicated")
    for relative in result.values():
        pure = PurePosixPath(relative)
        if pure.is_absolute() or pure.as_posix() != relative or ".." in pure.parts:
            raise ConcurrentRobustnessV2ReportError("report download path is unsafe")
    return result


def _downloads_html(downloads: Mapping[str, str]) -> str:
    links = "".join(
        f'<a download href="{html.escape(relative, quote=True)}"><strong>{html.escape(key)}</strong><code>{html.escape(relative)}</code></a>'
        for key, relative in downloads.items()
    )
    return (
        '<section class="robustness-v2-block" data-testid="robustness-v2-downloads">'
        f'<h3>{_bilingual("Manifest-bound downloads", "Manifest-bound downloads")}</h3>'
        f'<div class="robustness-v2-downloads">{links}</div></section>'
    )


_V2_CSS = r"""
.robustness-v2-report{--v2-ink:#172033;--v2-muted:#586273;--v2-line:#d7dee8;--v2-blue:#153f75;--v2-surface:#f4f7fb;color:var(--v2-ink);background:#fff;padding:clamp(2rem,4vw,4rem);border:2px solid var(--v2-blue);margin:1rem auto 5rem;max-width:1500px;font-family:inherit}.robustness-v2-report [hidden]{display:none!important}.robustness-v2-report h2{font-size:clamp(2rem,4vw,4.5rem);line-height:1.02;letter-spacing:-.045em;max-width:18ch;margin:.5rem 0 1rem}.robustness-v2-report p{line-height:1.6;color:var(--v2-muted);max-width:82ch}.robustness-v2-status{padding:.7rem 1rem;background:var(--v2-surface);border-left:4px solid var(--v2-blue)}.robustness-v2-controls{display:flex;flex-wrap:wrap;gap:.75rem;align-items:end;margin:2rem 0}.robustness-v2-controls label{display:grid;gap:.35rem;font-size:.78rem;font-weight:700}.robustness-v2-controls select,.robustness-v2-controls button{font:inherit;padding:.65rem .8rem;border:1px solid #9aa9bb;background:#fff;color:var(--v2-ink);border-radius:.25rem}.robustness-v2-controls button[aria-selected=true]{background:var(--v2-blue);color:#fff}.robustness-v2-controls :focus-visible,.robustness-v2-report a:focus-visible,.robustness-v2-report summary:focus-visible{outline:3px solid #5ea4ed;outline-offset:3px}.robustness-v2-block{padding:3rem 0;border-top:1px solid var(--v2-line)}.robustness-v2-block>h3{font-size:clamp(1.5rem,2.6vw,2.7rem);margin:0 0 1rem}.robustness-v2-segment{margin:1.5rem 0}.robustness-v2-segment h5{font-size:1.2rem;margin:.5rem 0}.robustness-v2-table-wrap{max-width:100%;overflow:auto;border:1px solid var(--v2-line)}.robustness-v2-table-wrap table{border-collapse:collapse;width:max-content;min-width:100%;font-size:.78rem}.robustness-v2-table-wrap th,.robustness-v2-table-wrap td{padding:.65rem .75rem;border-bottom:1px solid var(--v2-line);text-align:left;vertical-align:top;white-space:nowrap}.robustness-v2-table-wrap th{background:#eaf0f8;position:sticky;top:0}.robustness-v2-prompt-grid{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:1rem}.robustness-v2-prompt{min-width:0;border-top:4px solid var(--v2-blue);background:var(--v2-surface);padding:1rem}.robustness-v2-prompt pre,.robustness-v2-report details pre{white-space:pre-wrap;overflow-wrap:anywhere;max-height:26rem;overflow:auto;background:#172033;color:#f7fafc;padding:1rem;font:12px/1.55 ui-monospace,monospace}.robustness-v2-gateway-limit{margin-top:1rem;padding:1rem;background:#fff7e6;border-left:4px solid #b45309}.robustness-v2-mechanism-scroll{overflow:auto;border:1px solid var(--v2-line)}.robustness-v2-mechanism-scroll svg{display:block;width:100%;min-width:1200px;height:auto}.robustness-v2-downloads{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:.6rem}.robustness-v2-downloads a{display:grid;gap:.3rem;padding:.8rem;border:1px solid var(--v2-line);text-decoration:none;color:var(--v2-ink)}.robustness-v2-downloads code{overflow-wrap:anywhere;color:var(--v2-muted)}.robustness-v2-history{padding:1rem;background:#f7f4ec;border-left:4px solid #7a5a22}.robustness-v2-language{margin-left:auto}.robustness-v2-report[data-v2-state=error] [data-v2-interactive]{display:none!important}.robustness-v2-report[data-v2-state=error] .robustness-v2-status{border-color:#b42318;background:#fff1f0;color:#7a271a}@media(max-width:900px){.robustness-v2-prompt-grid,.robustness-v2-downloads{grid-template-columns:1fr}.robustness-v2-language{margin-left:0}}@media(max-width:640px){.robustness-v2-report{padding:1rem;border-left:0;border-right:0}.robustness-v2-controls{display:grid}.robustness-v2-controls label,.robustness-v2-controls select,.robustness-v2-controls button{width:100%}}
"""


_V2_SCRIPT = r"""
(() => {
  const root = document.querySelector('[data-testid="robustness-v2-report"]');
  if (!root) return;
  const status = root.querySelector('[data-testid="robustness-v2-state"]');
  const controls = Array.from(root.querySelectorAll('button,select'));
  const fail = (message) => {
    root.dataset.v2State = 'error';
    controls.forEach((control) => { control.disabled = true; });
    root.querySelectorAll('[data-v2-view]').forEach((view) => { view.hidden = true; });
    if (status) {
      status.dataset.v2State = 'error';
      status.textContent = message;
    }
  };
  try {
    if (window.__ABM_ROBUSTNESS_V2_FORCE_ERROR__) throw new Error('forced validation failure');
    const model = root.querySelector('[data-testid="robustness-v2-model-select"]');
    const prompt = root.querySelector('[data-testid="robustness-v2-prompt-select"]');
    const tabs = Array.from(root.querySelectorAll('[data-v2-view-button]'));
    const views = Array.from(root.querySelectorAll('[data-v2-view]'));
    if (!model || !prompt || tabs.length !== 2 || views.length !== 2) throw new Error('report controls incomplete');
    const applyFilters = () => {
      root.querySelectorAll('[data-v2-result-panel],[data-v2-judgment-panel]').forEach((panel) => {
        panel.hidden = panel.dataset.v2Model !== model.value || panel.dataset.v2Prompt !== prompt.value;
      });
    };
    const setView = (viewName) => {
      if (!['realized', 'judgment'].includes(viewName)) throw new Error('unsupported report view');
      root.dataset.v2ActiveView = viewName;
      tabs.forEach((tab) => tab.setAttribute('aria-selected', String(tab.dataset.v2ViewButton === viewName)));
      views.forEach((view) => { view.hidden = view.dataset.v2View !== viewName; });
    };
    const setLanguage = (language) => {
      if (!['zh-CN', 'en-US'].includes(language)) return;
      root.dataset.v2Language = language;
      root.querySelectorAll('[data-v2-lang]').forEach((node) => {
        node.hidden = node.dataset.v2Lang !== language;
      });
      root.querySelectorAll('[data-v2-language-button]').forEach((button) => {
        button.setAttribute('aria-pressed', String(button.dataset.v2LanguageButton === language));
      });
    };
    model.addEventListener('change', applyFilters);
    prompt.addEventListener('change', applyFilters);
    tabs.forEach((tab) => tab.addEventListener('click', () => setView(tab.dataset.v2ViewButton)));
    root.querySelectorAll('[data-v2-language-button]').forEach((button) => {
      button.addEventListener('click', () => setLanguage(button.dataset.v2LanguageButton));
    });
    document.querySelectorAll('[data-report-language]').forEach((button) => {
      button.addEventListener('click', () => setLanguage(button.dataset.reportLanguage));
    });
    applyFilters();
    setView('realized');
    setLanguage('zh-CN');
    root.dataset.v2State = 'ready';
    if (status) {
      status.dataset.v2State = 'ready';
      status.textContent = 'Validated projection ready · 已验证投影就绪';
    }
  } catch (error) {
    fail('Report validation failed closed · 报告验证失败并已关闭交互');
  }
})();
"""


def _report_section(
    projection: _ValidatedReportProjection,
    downloads: Mapping[str, str],
) -> str:
    mechanism = _MECHANISM_PRESENTATION.build_robustness_v2_master()
    if mechanism.semantic_set_identity_sha256 != projection.mechanism_identity_sha256:
        raise ConcurrentRobustnessV2ReportError("mechanism semantic identity changed during report rendering")
    zh_svg = _MECHANISM_PRESENTATION.render_inline_svg(mechanism, language="zh-CN")
    en_svg = _MECHANISM_PRESENTATION.render_inline_svg(mechanism, language="en-US")
    zh_fallback = _MECHANISM_PRESENTATION.render_fallback(mechanism, language="zh-CN")
    en_fallback = _MECHANISM_PRESENTATION.render_fallback(mechanism, language="en-US")
    model_options = "".join(
        f'<option value="{html.escape(model, quote=True)}">{html.escape(model)}</option>' for model in _V2_MODELS
    )
    prompt_options = "".join(f'<option value="{prompt}">{prompt}</option>' for prompt in _PROMPTS)
    section = (
        '<section class="robustness-v2-report" data-testid="robustness-v2-report" '
        'data-v2-state="loading" data-v2-active-view="realized" data-v2-language="zh-CN" '
        'data-production-deploy-eligible="false" data-provider-calls-during-composition="0" '
        'data-canonical-deployment-triggered="false">'
        f'<style>{_V2_CSS}</style>'
        '<p class="robustness-v2-kicker">Prompt–Model · Two-Stage Realized · v2</p>'
        f'<h2>{_bilingual("五模型 Prompt–Model Realized 主结果", "Five-model Prompt–Model Realized results")}</h2>'
        f'<p>{_bilingual("默认表格只展示 ABM Realized 互动；Provider Judgment 位于独立审计视图。固定样本、固定互动图、共享 draw，每个 cell 只有一条 realized path。", "The default tables show ABM Realized interactions only; Provider Judgment stays in a separate audit view. The sample and graph are fixed, the draw is shared, and each cell has one realized path.")}</p>'
        '<p class="robustness-v2-status" data-testid="robustness-v2-state" data-v2-state="loading" aria-live="polite">'
        'Validating projection · 正在验证投影</p>'
        '<div class="robustness-v2-controls" data-v2-interactive>'
        '<button type="button" data-v2-view-button="realized" aria-selected="true">Realized</button>'
        '<button type="button" data-v2-view-button="judgment" aria-selected="false">Judgment Audit</button>'
        f'<label>{_bilingual("模型", "Model")}<select data-testid="robustness-v2-model-select">{model_options}</select></label>'
        f'<label>{_bilingual("Prompt", "Prompt")}<select data-testid="robustness-v2-prompt-select">{prompt_options}</select></label>'
        '<div class="robustness-v2-language" role="group" aria-label="Language">'
        '<button type="button" data-v2-language-button="zh-CN" aria-pressed="true">中文</button>'
        '<button type="button" data-v2-language-button="en-US" aria-pressed="false">English</button></div></div>'
        '<section class="robustness-v2-block" data-v2-view="realized" data-testid="robustness-v2-realized-view" data-v2-interactive>'
        f'<h3>{_bilingual("Realized Main：S1–S3 × M1–M3", "Realized Main: S1–S3 × M1–M3")}</h3>'
        f'{_realized_panels(projection)}</section>'
        '<section class="robustness-v2-block" data-v2-view="judgment" data-testid="robustness-v2-judgment-view" data-v2-interactive hidden>'
        f'<h3>{_bilingual("Judgment Audit：意向不是已实现行动", "Judgment Audit: intent is not a realized action")}</h3>'
        f'{_judgment_panels(projection)}</section>'
        f'{_prompt_catalog_html(projection)}{_provider_audit_html(projection)}'
        '<section class="robustness-v2-block" data-testid="robustness-v2-mechanism">'
        f'<h3>{_bilingual("Judgment → Realization → Barrier → Feedback", "Judgment → Realization → Barrier → Feedback")}</h3>'
        f'<div class="robustness-v2-mechanism-scroll" data-v2-lang="zh-CN">{zh_svg}</div>'
        f'<div class="robustness-v2-mechanism-scroll" data-v2-lang="en-US" hidden>{en_svg}</div>'
        f'<div data-v2-lang="zh-CN">{zh_fallback}</div><div data-v2-lang="en-US" hidden>{en_fallback}</div></section>'
        '<section class="robustness-v2-block"><div class="robustness-v2-history" '
        'data-testid="robustness-v2-historical-reference"><strong>Historical 16-cell Judgment Reference</strong>'
        '<p>Historical OpenAI 4 × 4 cells remain immutable Judgment-era evidence and do not enter the new Realized Main denominator. '
        '历史 16 cells 与六张机制图保持独立、原字节和原 hash。</p><a href="#historical-sensitivity-1000">Open Historical reference</a></div></section>'
        f'{_downloads_html(downloads)}</section>'
    )
    return section


def _render_report_html(
    base_report: bytes,
    projection: _ValidatedReportProjection,
    downloads: Mapping[str, str],
) -> bytes:
    try:
        report = base_report.decode("utf-8")
    except UnicodeDecodeError as exc:
        raise ConcurrentRobustnessV2ReportError("base Full-Pool report is not UTF-8") from exc
    main_matches = list(re.finditer(r'<main class="full-pool-presentation"[^>]*>', report))
    if len(main_matches) != 1 or report.count("</body>") != 1:
        raise ConcurrentRobustnessV2ReportError("base Full-Pool report root is missing or duplicated")
    section = _report_section(projection, downloads)
    match = main_matches[0]
    report = report[: match.end()] + section + report[match.end() :]
    report = report.replace("</body>", f"<script>{_V2_SCRIPT}</script></body>")
    payload = report.encode("utf-8")
    if len(payload) >= _MAX_HTML_BYTES:
        raise ConcurrentRobustnessV2ReportError("v2 report exceeds the 3 MiB presentation limit")
    if payload.count(b'<section class="robustness-v2-report" data-testid="robustness-v2-report"') != 1:
        raise ConcurrentRobustnessV2ReportError("v2 report root marker is crossed")
    if b"realized_reason" in payload:
        raise ConcurrentRobustnessV2ReportError("v2 report invents a Realized reason")
    return payload


def _prompt_catalog_document(projection: _ValidatedReportProjection) -> dict[str, Any]:
    return {
        "schema_version": _PROMPT_CATALOG_SCHEMA,
        "catalog_scope": "complete_static_client_submitted_templates_only",
        "per_user_rendered_prompts_persisted": False,
        "prompts": [dict(record) for record in projection.prompt_catalog],
        "messages": [dict(record) for record in projection.message_catalog],
        "gateway_limitation": {
            "route": "Antigravity OpenAI-compatible gateway",
            "direct_gemini_developer_api": False,
            "unobservable_gateway_context_possible": True,
            "client_submitted_prompt_is_complete_effective_prompt": False,
        },
    }


def _new_artifact_payloads(projection: _ValidatedReportProjection) -> dict[str, bytes]:
    projection_document = projection.document()
    projection_bytes = _canonical_json_bytes(projection_document)
    mechanism = _MECHANISM_PRESENTATION.build_robustness_v2_master()
    mermaid = mechanism.mermaid_artifacts[0]
    if mermaid.filename != _MECHANISM_MMD or mermaid.sha256 != _sha256_bytes(mermaid.payload):
        raise ConcurrentRobustnessV2ReportError("v2 mechanism artifact is crossed")
    workbook = _workbook_bytes(projection)
    if workbook != _workbook_bytes(projection):
        raise ConcurrentRobustnessV2ReportError("repeated workbook builds differ")
    _validate_workbook(workbook, projection)
    return {
        _PROJECTION_JSON: projection_bytes,
        _REALIZED_CSV: _csv_bytes(_REALIZED_FIELDS, projection.realized_main_rows),
        _JUDGMENT_CSV: _csv_bytes(_JUDGMENT_FIELDS, projection.judgment_audit_rows),
        _PROMPT_CATALOG_JSON: _canonical_json_bytes(_prompt_catalog_document(projection)),
        _PROVIDER_CSV: _csv_bytes(_PROVIDER_FIELDS, projection.provider_audit_rows),
        _CELL_BATCH_CSV: _csv_bytes(_CELL_BATCH_FIELDS, projection.cell_batch_evidence_rows),
        _WORKBOOK: workbook,
        _MECHANISM_MMD: mermaid.payload,
    }


def _report_payload_document(
    projection: _ValidatedReportProjection,
    *,
    downloads: Mapping[str, str],
) -> dict[str, Any]:
    projection_bytes = _canonical_json_bytes(projection.document())
    return {
        "schema_version": _REPORT_SCHEMA,
        "candidate_type": _CANDIDATE_TYPE,
        "title": "Prompt–Model Realized Robustness · Table-First Teacher Report",
        "primary_outcome": "abm_realized_engagement",
        "source_lineage": dict(projection.source_lineage),
        "source_lineage_identity_sha256": _sha256_bytes(
            _canonical_json_bytes(dict(projection.source_lineage))
        ),
        "projection_schema_version": _PROJECTION_SCHEMA,
        "projection_sha256": _sha256_bytes(projection_bytes),
        "mechanism_schema_version": projection.mechanism_schema_version,
        "mechanism_identity_sha256": projection.mechanism_identity_sha256,
        "approved_downloads": dict(downloads),
        "claim_boundary": dict(projection.claim_boundary),
        "provider_calls_during_composition": 0,
        "canonical_deployment_triggered": False,
        "production_deploy_eligible": False,
    }


def _manifest_document(
    projection: _ValidatedReportProjection,
    *,
    artifact_hashes: Mapping[str, str],
    artifact_sizes: Mapping[str, int],
    downloads: Mapping[str, str],
) -> dict[str, Any]:
    records = [
        {
            "relative_path": path,
            "sha256": artifact_hashes[path],
            "bytes": artifact_sizes[path],
        }
        for path in sorted(artifact_hashes)
    ]
    identity = _sha256_bytes(_canonical_json_bytes(dict(sorted(artifact_hashes.items()))))
    return {
        "schema_version": _CANDIDATE_MANIFEST_SCHEMA,
        "candidate_type": _CANDIDATE_TYPE,
        "status": "complete_non_deployable_candidate",
        "candidate_identity_sha256": identity,
        "source_lineage": dict(projection.source_lineage),
        "source_lineage_identity_sha256": _sha256_bytes(
            _canonical_json_bytes(dict(projection.source_lineage))
        ),
        "projection_sha256": artifact_hashes[_PROJECTION_JSON],
        "report_sha256": artifact_hashes[_REPORT_HTML],
        "workbook_sha256": artifact_hashes[_WORKBOOK],
        "mechanism_identity_sha256": projection.mechanism_identity_sha256,
        "artifacts": records,
        "approved_downloads": dict(downloads),
        "provider_calls_during_composition": 0,
        "canonical_deployment_triggered": False,
        "production_deploy_eligible": False,
    }


def _expected_candidate(
    *,
    base_bundle: Path,
    projection: _ValidatedReportProjection,
) -> _ExpectedCandidate:
    base_hashes, base_sizes = _tree_records(base_bundle)
    if _REPORT_HTML not in base_hashes or _MANIFEST in base_hashes:
        raise ConcurrentRobustnessV2ReportError("base Full-Pool bundle inventory is crossed")
    new_payloads = _new_artifact_payloads(projection)
    collisions = set(new_payloads).intersection(base_hashes)
    if collisions:
        raise ConcurrentRobustnessV2ReportError("v2 report artifacts collide with the base bundle")
    base_report = (base_bundle / _REPORT_HTML).read_bytes()
    downloads = _download_mapping(base_report.decode("utf-8"))
    report_payload = _canonical_json_bytes(
        _report_payload_document(projection, downloads=downloads)
    )
    new_payloads[_REPORT_PAYLOAD] = report_payload
    final_report = _render_report_html(base_report, projection, downloads)
    artifact_hashes = dict(base_hashes)
    artifact_sizes = dict(base_sizes)
    artifact_hashes[_REPORT_HTML] = _sha256_bytes(final_report)
    artifact_sizes[_REPORT_HTML] = len(final_report)
    for path, payload in new_payloads.items():
        artifact_hashes[path] = _sha256_bytes(payload)
        artifact_sizes[path] = len(payload)
    for relative in downloads.values():
        if relative not in artifact_hashes:
            raise ConcurrentRobustnessV2ReportError(f"download is missing from candidate inventory: {relative}")
    manifest = _manifest_document(
        projection,
        artifact_hashes=artifact_hashes,
        artifact_sizes=artifact_sizes,
        downloads=downloads,
    )
    return _ExpectedCandidate(
        new_payloads=new_payloads,
        final_report=final_report,
        manifest=manifest,
        approved_downloads=downloads,
        artifact_hashes=artifact_hashes,
        artifact_sizes=artifact_sizes,
    )


def _base_bundle(
    *,
    source: ClosedFullPoolTwoStageSource,
    historical_candidate: Path,
    parent: Path,
) -> tuple[Path, Path]:
    temporary = Path(tempfile.mkdtemp(prefix=".robustness-v2-base.", dir=parent))
    bundle = temporary / "full-pool-bundle"
    try:
        historical_hashes, _ = _tree_records(historical_candidate)
        compose_full_pool_presentation_bundle(
            source=source,
            historical_candidate=historical_candidate,
            historical_inventory=historical_hashes,
            destination=bundle,
        )
        validate_full_pool_presentation_bundle(
            bundle,
            source=source,
            historical_candidate=historical_candidate,
        )
    except Exception:
        shutil.rmtree(temporary, ignore_errors=True)
        raise
    return temporary, bundle


def _validate_candidate_directory(
    candidate: Path,
    *,
    base_bundle: Path,
    projection: _ValidatedReportProjection,
) -> _V2RealizedCandidateFacts:
    if candidate.is_symlink() or not candidate.is_dir():
        raise ConcurrentRobustnessV2ReportError("v2 report candidate must be a real directory")
    expected = _expected_candidate(base_bundle=base_bundle, projection=projection)
    actual_hashes, actual_sizes = _tree_records(candidate)
    expected_paths = set(expected.artifact_hashes) | {_MANIFEST}
    if set(actual_hashes) != expected_paths:
        raise ConcurrentRobustnessV2ReportError("v2 report candidate has missing or extra artifacts")
    base_hashes, _ = _tree_records(base_bundle)
    for path, sha256 in base_hashes.items():
        if path != _REPORT_HTML and actual_hashes.get(path) != sha256:
            raise ConcurrentRobustnessV2ReportError("closed Full-Pool or Historical bytes changed in candidate")
    for path, payload in expected.new_payloads.items():
        if (candidate / path).read_bytes() != payload:
            raise ConcurrentRobustnessV2ReportError(f"v2 report artifact is not reproducible: {path}")
    if (candidate / _REPORT_HTML).read_bytes() != expected.final_report:
        raise ConcurrentRobustnessV2ReportError("v2 report HTML differs from the validated projection")
    manifest_bytes = _canonical_json_bytes(expected.manifest)
    if (candidate / _MANIFEST).read_bytes() != manifest_bytes:
        raise ConcurrentRobustnessV2ReportError("v2 candidate manifest is not reproducible")
    manifest = json.loads(manifest_bytes)
    records = _object_rows(manifest.get("artifacts"), "v2 candidate artifacts")
    if [str(row.get("relative_path")) for row in records] != sorted(expected.artifact_hashes):
        raise ConcurrentRobustnessV2ReportError("v2 candidate artifact inventory is non-canonical")
    for row in records:
        relative = str(row["relative_path"])
        if (
            row.get("sha256") != actual_hashes[relative]
            or row.get("bytes") != actual_sizes[relative]
        ):
            raise ConcurrentRobustnessV2ReportError("v2 candidate artifact hash or size is crossed")
    report = expected.final_report.decode("utf-8")
    observed_downloads = {
        html.unescape(match)
        for match in re.findall(
            r'<a\b(?=[^>]*\bdownload\b)[^>]*\bhref="([^"]+)"',
            report,
            re.IGNORECASE,
        )
    }
    if observed_downloads != set(expected.approved_downloads.values()):
        raise ConcurrentRobustnessV2ReportError("v2 report links differ from approved downloads")
    _validate_workbook((candidate / _WORKBOOK).read_bytes(), projection)
    prompt_catalog = json.loads((candidate / _PROMPT_CATALOG_JSON).read_text(encoding="utf-8"))
    if (
        prompt_catalog.get("schema_version") != _PROMPT_CATALOG_SCHEMA
        or prompt_catalog.get("per_user_rendered_prompts_persisted") is not False
        or len(prompt_catalog.get("prompts", ())) != 4
        or len(prompt_catalog.get("messages", ())) != 3
    ):
        raise ConcurrentRobustnessV2ReportError("Prompt Catalog contract is crossed")
    if "realized_reason" in report.lower():
        raise ConcurrentRobustnessV2ReportError("v2 report contains an invented Realized field")
    return _V2RealizedCandidateFacts(
        root=candidate.resolve(strict=True),
        manifest_sha256=actual_hashes[_MANIFEST],
        candidate_identity_sha256=str(manifest["candidate_identity_sha256"]),
        projection_sha256=actual_hashes[_PROJECTION_JSON],
        report_sha256=actual_hashes[_REPORT_HTML],
        workbook_sha256=actual_hashes[_WORKBOOK],
        artifact_hashes={path: actual_hashes[path] for path in sorted(expected.artifact_hashes)},
        approved_downloads=dict(expected.approved_downloads),
    )


def compose_v2_realized_candidate(
    *,
    source: ClosedFullPoolTwoStageSource,
    historical_candidate: Path,
    v2_source: _ConcurrentRobustnessV2ReportSource,
    source_lineage: Mapping[str, Any],
    destination: Path,
) -> Path:
    """Build one atomic nondeployable candidate from independently closed facts."""

    projection = _build_projection(source=v2_source, source_lineage=source_lineage)
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary, base = _base_bundle(
        source=source,
        historical_candidate=historical_candidate,
        parent=destination.parent,
    )
    staging = Path(
        tempfile.mkdtemp(
            prefix=f".{destination.name}.robustness-v2.",
            suffix=".staging",
            dir=destination.parent,
        )
    )
    installed = False
    try:
        shutil.copytree(base, staging, dirs_exist_ok=True, copy_function=shutil.copyfile)
        expected = _expected_candidate(base_bundle=base, projection=projection)
        (staging / _REPORT_HTML).write_bytes(expected.final_report)
        for relative, payload in expected.new_payloads.items():
            target = staging / relative
            if target.exists() or target.is_symlink():
                raise ConcurrentRobustnessV2ReportError("v2 report attempted to overwrite a closed artifact")
            target.write_bytes(payload)
        (staging / _MANIFEST).write_bytes(_canonical_json_bytes(expected.manifest))
        _validate_candidate_directory(staging, base_bundle=base, projection=projection)
        if os.path.lexists(destination):
            raise ConcurrentRobustnessV2ReportError("v2 report candidate destination appeared during publication")
        os.replace(staging, destination)
        installed = True
        _validate_candidate_directory(destination, base_bundle=base, projection=projection)
        _assert_concurrent_robustness_v2_report_source_unchanged(v2_source)
        installed = False
        return destination.resolve(strict=True)
    except Exception:
        if staging.is_dir() and not staging.is_symlink():
            shutil.rmtree(staging, ignore_errors=True)
        if installed and destination.is_dir() and not destination.is_symlink():
            shutil.rmtree(destination, ignore_errors=True)
        raise
    finally:
        shutil.rmtree(temporary, ignore_errors=True)


def validate_v2_realized_candidate(
    candidate: Path,
    *,
    source: ClosedFullPoolTwoStageSource,
    historical_candidate: Path,
    v2_source: _ConcurrentRobustnessV2ReportSource,
    source_lineage: Mapping[str, Any],
) -> _V2RealizedCandidateFacts:
    """Independently rebuild projection, workbook, HTML, and manifest from closed inputs."""

    projection = _build_projection(source=v2_source, source_lineage=source_lineage)
    temporary, base = _base_bundle(
        source=source,
        historical_candidate=historical_candidate,
        parent=candidate.parent,
    )
    try:
        facts = _validate_candidate_directory(candidate, base_bundle=base, projection=projection)
        _assert_concurrent_robustness_v2_report_source_unchanged(v2_source)
        return facts
    finally:
        shutil.rmtree(temporary, ignore_errors=True)


__all__: list[str] = []
